from pathlib import Path

import pandas as pd
import pytest

from converter import (
    ConversionOptions,
    dataframe_to_latex,
    parse_scale_factor,
    read_table_file,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _simple_df() -> pd.DataFrame:
    """2x2 DataFrame for quick tests."""
    return pd.DataFrame({"A": [1, 3], "B": [2, 4]})


def _three_col_df() -> pd.DataFrame:
    return pd.DataFrame(
        [["alpha", "beta", "gamma"]],
        columns=["First", "Second", "Third"],
    )


# ---------------------------------------------------------------------------
# Updated existing tests
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_2x2_dataframe_contains_default_table_parts() -> None:
    # Arrange
    dataframe = _simple_df()

    # Act
    latex_code = dataframe_to_latex(dataframe)

    # Assert — defaults: all borders, float [htbp]
    assert r"\begin{table}[htbp]" in latex_code
    assert r"\begin{tabular}{|c|c|}" in latex_code
    assert r"A & B \\" in latex_code
    assert r"1 & 2 \\" in latex_code
    assert r"3 & 4 \\" in latex_code


def test_dataframe_to_latex_with_empty_dataframe_raises_clear_error() -> None:
    # Arrange
    dataframe = pd.DataFrame()

    # Act / Assert
    with pytest.raises(ValueError, match="Input table is empty"):
        dataframe_to_latex(dataframe)


def test_dataframe_to_latex_with_three_columns_uses_matching_structure() -> None:
    # Arrange
    dataframe = _three_col_df()

    # Act
    latex_code = dataframe_to_latex(dataframe)
    latex_lines = latex_code.splitlines()

    # Assert — default options: all-border grid, float [htbp]
    assert latex_lines == [
        r"\begin{table}[htbp]",
        r"\centering",
        r"\begin{tabular}{|c|c|c|}",
        r"\hline",
        r"First & Second & Third \\",
        r"\hline",
        r"alpha & beta & gamma \\",
        r"\hline",
        r"\end{tabular}",
        r"\end{table}",
    ]


def test_dataframe_to_latex_with_caption_and_label_includes_both_commands() -> None:
    # Arrange
    dataframe = pd.DataFrame({"A": [1], "B": [2]})
    options = ConversionOptions(caption="実験1", label="tab:exp1")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\caption{実験1}" in latex_code
    assert r"\label{tab:exp1}" in latex_code


def test_read_table_file_with_csv_path_returns_dataframe(tmp_path: Path) -> None:
    # Arrange
    csv_path = tmp_path / "sample.csv"
    csv_path.write_text("A,B\n1,2\n3,4\n", encoding="utf-8")

    # Act
    dataframe = read_table_file(csv_path)

    # Assert
    assert dataframe.to_dict(orient="list") == {"A": [1, 3], "B": [2, 4]}


# ---------------------------------------------------------------------------
# Default output structure
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_default_options_produces_full_grid() -> None:
    # Arrange
    dataframe = _simple_df()

    # Act
    latex_code = dataframe_to_latex(dataframe)
    lines = latex_code.splitlines()

    # Assert — full grid: vertical bars, \hline everywhere
    assert lines == [
        r"\begin{table}[htbp]",
        r"\centering",
        r"\begin{tabular}{|c|c|}",
        r"\hline",
        r"A & B \\",
        r"\hline",
        r"1 & 2 \\",
        r"\hline",
        r"3 & 4 \\",
        r"\hline",
        r"\end{tabular}",
        r"\end{table}",
    ]


# ---------------------------------------------------------------------------
# Escaping
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_escape_ampersand_in_cell_escapes_correctly() -> None:
    # Arrange
    dataframe = pd.DataFrame({"Col": ["A & B"]})

    # Act
    latex_code = dataframe_to_latex(dataframe)

    # Assert
    assert r"A \& B" in latex_code


def test_dataframe_to_latex_with_escape_underscore_in_cell_escapes_correctly() -> None:
    # Arrange
    dataframe = pd.DataFrame({"Col": ["result_1"]})

    # Act
    latex_code = dataframe_to_latex(dataframe)

    # Assert
    assert r"result\_1" in latex_code


def test_dataframe_to_latex_with_escape_backslash_in_cell_escapes_correctly() -> None:
    # Arrange
    dataframe = pd.DataFrame({"Col": ["a\\b"]})

    # Act
    latex_code = dataframe_to_latex(dataframe)

    # Assert
    assert r"a\textbackslash{}b" in latex_code


def test_dataframe_to_latex_with_escape_caption_text_escapes_special_chars() -> None:
    # Arrange
    dataframe = pd.DataFrame({"A": [1]})
    options = ConversionOptions(caption="100% done & finished")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\caption{100\% done \& finished}" in latex_code


def test_dataframe_to_latex_with_escape_disabled_leaves_chars_raw() -> None:
    # Arrange
    dataframe = pd.DataFrame({"Col": ["A & B"]})
    options = ConversionOptions(escape=False)

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert — raw ampersand, no backslash-escape
    assert "A & B" in latex_code
    assert r"\&" not in latex_code


def test_dataframe_to_latex_with_escape_label_not_escaped() -> None:
    # Arrange
    dataframe = pd.DataFrame({"A": [1]})
    options = ConversionOptions(label="tab:my_table")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert — underscore in label must NOT be escaped
    assert r"\label{tab:my_table}" in latex_code


# ---------------------------------------------------------------------------
# Bold
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_bold_first_row_wraps_headers() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(bold_first_row=True)

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\textbf{A} & \textbf{B} \\" in latex_code
    # Body rows must not be bold
    assert r"\textbf{1}" not in latex_code


def test_dataframe_to_latex_with_bold_first_column_wraps_first_body_cells() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(bold_first_column=True)

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\textbf{1} & 2 \\" in latex_code
    assert r"\textbf{3} & 4 \\" in latex_code
    # Header first cell should NOT be bold (bold_first_row is False)
    assert r"\textbf{A}" not in latex_code


def test_dataframe_to_latex_with_both_bold_options_no_double_wrap() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(bold_first_row=True, bold_first_column=True)

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert — header first cell bold from bold_first_row, not double-wrapped
    assert r"\textbf{A} & \textbf{B} \\" in latex_code
    assert r"\textbf{\textbf{" not in latex_code
    # Body first column bold
    assert r"\textbf{1} & 2 \\" in latex_code


# ---------------------------------------------------------------------------
# Border styles
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_booktabs_border_produces_correct_rules() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(border_style="booktabs")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\toprule" in latex_code
    assert r"\midrule" in latex_code
    assert r"\bottomrule" in latex_code
    assert r"\hline" not in latex_code
    # No vertical bars
    assert r"\begin{tabular}{cc}" in latex_code


def test_dataframe_to_latex_with_none_border_produces_no_rules() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(border_style="none")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\hline" not in latex_code
    assert r"\toprule" not in latex_code
    assert r"\midrule" not in latex_code
    assert r"\bottomrule" not in latex_code
    assert r"\begin{tabular}{cc}" in latex_code


def test_dataframe_to_latex_with_horizontal_border_produces_three_hlines() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(border_style="horizontal")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)
    lines = latex_code.splitlines()

    # Assert — exactly 3 \hline: top, after header, bottom
    hline_count = sum(1 for line in lines if line.strip() == r"\hline")
    assert hline_count == 3
    # No vertical bars
    assert r"\begin{tabular}{cc}" in latex_code


# ---------------------------------------------------------------------------
# Table types
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_longtable_type_no_table_wrapper() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(table_type="longtable")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\begin{longtable}" in latex_code
    assert r"\end{longtable}" in latex_code
    assert r"\begin{table}" not in latex_code
    assert r"\end{table}" not in latex_code


def test_dataframe_to_latex_with_longtable_caption_inside() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(
        table_type="longtable",
        caption="My Table",
        label="tab:my",
    )

    # Act
    latex_code = dataframe_to_latex(dataframe, options)
    lines = latex_code.splitlines()

    # Assert — caption and label appear inside longtable, before rules
    longtable_idx = next(
        i for i, line in enumerate(lines) if r"\begin{longtable}" in line
    )
    caption_idx = next(i for i, line in enumerate(lines) if r"\caption{" in line)
    label_idx = next(i for i, line in enumerate(lines) if r"\label{" in line)
    assert caption_idx > longtable_idx
    assert label_idx > caption_idx
    # Caption line ends with \\
    assert lines[caption_idx].endswith("\\\\")


def test_dataframe_to_latex_with_tabularx_type_uses_textwidth() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(table_type="tabularx")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\begin{tabularx}{\textwidth}" in latex_code
    assert r"\end{tabularx}" in latex_code
    # Still inside a table float
    assert r"\begin{table}" in latex_code


# ---------------------------------------------------------------------------
# Scale box
# ---------------------------------------------------------------------------


def test_parse_scale_factor_with_blank_value_disables_scaling() -> None:
    assert parse_scale_factor("") is None
    assert parse_scale_factor("   ") is None


def test_parse_scale_factor_with_positive_number_returns_float() -> None:
    assert parse_scale_factor("0.85") == 0.85


@pytest.mark.parametrize("value", ["0", "-1", "not-a-number", "nan", "inf"])
def test_parse_scale_factor_with_invalid_value_raises(value: str) -> None:
    with pytest.raises(ValueError, match="positive number"):
        parse_scale_factor(value)


def test_dataframe_to_latex_with_scale_factor_wraps_tabular() -> None:
    latex_code = dataframe_to_latex(_simple_df(), ConversionOptions(scale_factor=0.85))

    assert r"\scalebox{0.85}{%" in latex_code
    assert latex_code.index(r"\scalebox{0.85}{%") < latex_code.index(r"\begin{tabular}")
    assert latex_code.index(r"\end{tabular}") < latex_code.rindex("}")


def test_dataframe_to_latex_with_scale_factor_full_document_adds_graphicx() -> None:
    latex_code = dataframe_to_latex(
        _simple_df(),
        ConversionOptions(scale_factor=0.9, full_document=True),
    )

    assert r"\usepackage{graphicx}" in latex_code


def test_dataframe_to_latex_with_longtable_does_not_apply_scale_box() -> None:
    latex_code = dataframe_to_latex(
        _simple_df(),
        ConversionOptions(scale_factor=0.9, table_type="longtable"),
    )

    assert r"\scalebox" not in latex_code


# ---------------------------------------------------------------------------
# Text alignment
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_left_alignment_changes_colspec() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(text_alignment="l", border_style="none")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\begin{tabular}{ll}" in latex_code


def test_dataframe_to_latex_with_right_alignment_changes_colspec() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(text_alignment="r", border_style="none")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\begin{tabular}{rr}" in latex_code


# ---------------------------------------------------------------------------
# Table alignment
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_left_table_alignment_produces_raggedright() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(table_alignment="left")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\raggedright" in latex_code
    assert r"\centering" not in latex_code


def test_dataframe_to_latex_with_right_table_alignment_produces_raggedleft() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(table_alignment="right")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\raggedleft" in latex_code


# ---------------------------------------------------------------------------
# Full document (MWE)
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_full_document_wraps_with_documentclass() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(full_document=True)

    # Act
    latex_code = dataframe_to_latex(dataframe, options)
    lines = latex_code.splitlines()

    # Assert
    assert lines[0] == r"\documentclass{article}"
    assert r"\begin{document}" in latex_code
    assert r"\end{document}" in latex_code
    assert lines[-1] == r"\end{document}"


def test_dataframe_to_latex_with_full_document_booktabs_includes_package() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(full_document=True, border_style="booktabs")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\usepackage{booktabs}" in latex_code


def test_dataframe_to_latex_with_full_document_longtable_includes_package() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(full_document=True, table_type="longtable")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\usepackage{longtable}" in latex_code


def test_dataframe_to_latex_with_full_document_tabularx_includes_package() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(full_document=True, table_type="tabularx")

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\usepackage{tabularx}" in latex_code


def test_dataframe_to_latex_with_full_document_no_extra_packages_when_not_needed() -> (
    None
):
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(full_document=True)

    # Act
    latex_code = dataframe_to_latex(dataframe, options)

    # Assert
    assert r"\usepackage" not in latex_code


# ---------------------------------------------------------------------------
# Float position
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_float_position_disabled_no_brackets() -> None:
    # Arrange
    dataframe = _simple_df()
    options = ConversionOptions(use_float_position=False)

    # Act
    latex_code = dataframe_to_latex(dataframe, options)
    lines = latex_code.splitlines()

    # Assert — \begin{table} without [htbp]
    table_line = next(line for line in lines if line.startswith(r"\begin{table}"))
    assert table_line == r"\begin{table}"


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


def test_dataframe_to_latex_with_empty_dataframe_still_raises_value_error() -> None:
    # Arrange
    dataframe = pd.DataFrame({"A": []})

    # Act / Assert
    with pytest.raises(ValueError, match="Input table is empty"):
        dataframe_to_latex(dataframe)


# ---------------------------------------------------------------------------
# Leading empty rows / columns (read_table_file bug fix)
# ---------------------------------------------------------------------------


def test_read_table_file_with_xlsx_leading_empties_skips_to_real_data(
    tmp_path: Path,
) -> None:
    """XLSX with an empty first row and empty first column is trimmed correctly.

    Grid layout (1-indexed cells):
        A1:  ""    B1: ""     C1: ""           D1: ""
        A2:  ""    B2: "国"   C2: "鶏肉世界比 %"  D2: "鶏卵世界比 %"
        A3:  ""    B3: "アメリカ" C3: 15.7        D3: 7.2
        A4:  ""    B4: "日本"   C4: 1.9          D4: 2.7
    """
    import openpyxl

    # Arrange
    wb = openpyxl.Workbook()
    ws = wb.active
    # Row 1: all empty (left as default None)
    # Row 2 onward: col A empty, cols B-D have data
    ws["B2"] = "国"
    ws["C2"] = "鶏肉世界比 %"
    ws["D2"] = "鶏卵世界比 %"
    ws["B3"] = "アメリカ"
    ws["C3"] = 15.7
    ws["D3"] = 7.2
    ws["B4"] = "日本"
    ws["C4"] = 1.9
    ws["D4"] = 2.7
    xlsx_path = tmp_path / "leading_empties.xlsx"
    wb.save(xlsx_path)

    # Act
    dataframe = read_table_file(xlsx_path)

    # Assert
    assert list(dataframe.columns) == ["国", "鶏肉世界比 %", "鶏卵世界比 %"]
    first_row = list(dataframe.iloc[0])
    assert first_row == ["アメリカ", 15.7, 7.2]
    assert len(dataframe) == 2


def test_read_table_file_with_csv_leading_empties_skips_to_real_data(
    tmp_path: Path,
) -> None:
    """CSV with an empty first line and leading empty column is trimmed."""
    # Arrange
    csv_content = (
        ",,,\n"  # row 0: all empty
        ",国,鶏肉世界比 %,鶏卵世界比 %\n"  # row 1: header (col 0 empty)
        ",アメリカ,15.7,7.2\n"
        ",日本,1.9,2.7\n"
    )
    csv_path = tmp_path / "leading_empties.csv"
    csv_path.write_text(csv_content, encoding="utf-8")

    # Act
    dataframe = read_table_file(csv_path)

    # Assert
    assert list(dataframe.columns) == ["国", "鶏肉世界比 %", "鶏卵世界比 %"]
    first_row = list(dataframe.iloc[0])
    assert first_row == ["アメリカ", 15.7, 7.2]
    assert len(dataframe) == 2


def test_read_table_file_with_clean_csv_preserves_int_dtypes(
    tmp_path: Path,
) -> None:
    """Clean CSV (no leading empties) still returns integer columns."""
    # Arrange
    csv_path = tmp_path / "clean.csv"
    csv_path.write_text("A,B\n1,2\n3,4\n", encoding="utf-8")

    # Act
    dataframe = read_table_file(csv_path)

    # Assert — must be identical to original behavior
    assert dataframe.to_dict(orient="list") == {"A": [1, 3], "B": [2, 4]}


def test_read_table_file_with_all_empty_csv_returns_empty_dataframe(
    tmp_path: Path,
) -> None:
    """A file that is entirely empty cells returns an empty DataFrame."""
    # Arrange
    csv_path = tmp_path / "empty.csv"
    csv_path.write_text(",,,\n,,,\n", encoding="utf-8")

    # Act
    dataframe = read_table_file(csv_path)

    # Assert
    assert dataframe.empty
